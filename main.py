# main.py — Email Filler (Simple UI)
# Hybrid: infer from uploaded file first, then fall back to persistent repository.
# Minimal interface: short repo check, progress bar, concise summary, CSV + Excel (highlighted).

import os, re, io
from collections import defaultdict, Counter

import pandas as pd
import streamlit as st
from unidecode import unidecode

st.set_page_config(page_title="Email Filler", layout="wide")

# -----------------------------
# Config / Repo paths
# -----------------------------
CANDIDATE_REPO_PATHS = [
    os.environ.get("REPO_PATH", ""),  # optional Railway env var
    "./master_repository.parquet",
    "./master_repository_compact.parquet",
    "/data/master_repository.parquet",
    "/data/master_repository_compact.parquet",
]
REPO_SCHEMA = ["company", "country_norm", "domain", "pattern", "count"]

LEGAL_STOPWORDS = {
    "the","group","holding","holdings","international","company","co","inc","incorporated",
    "limited","ltd","plc","llc","llp","gmbh","ag","sa","spa","srl","bv","nv","as","ab","oy","aps",
    "kft","zrt","rt","sarl","sas","pte","pty","bhd","sdn","kk","dmcc","pjsc","jsc","ltda",
    "corp","corporation","co."
}

# -----------------------------
# Normalizers
# -----------------------------
def normalize_company_key(name: str) -> str:
    if not isinstance(name, str) or not name.strip():
        return ""
    s = unidecode(name).lower()
    s = re.sub(r"[^a-z0-9\s-]", " ", s)
    toks = [t for t in re.split(r"[\s-]+", s) if t]
    toks = [t for t in toks if t not in LEGAL_STOPWORDS]
    return "".join(toks)

def normalize_country(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    s = unidecode(str(x)).strip().lower()
    if not s:
        return None
    alias = {
        "u.s.": "united states",
        "usa": "united states",
        "u.k.": "united kingdom",
        "uk": "united kingdom",
        "england": "united kingdom",
        "gb": "united kingdom",
        "great britain": "united kingdom",
    }
    return alias.get(s, s)

def norm_name(x):
    if pd.isna(x): return ""
    x = unidecode(str(x)).lower().strip()
    x = re.sub(r"[^\w\s'-]", "", x)
    x = x.replace("’","'")
    return re.sub(r"[\s'-]+","",x)

# -----------------------------
# Detect pattern from a known email + names
# -----------------------------
def detect_email_pattern(first, last, email):
    if not isinstance(email, str) or "@" not in email:
        return None
    local = email.split("@")[0].lower()
    f = norm_name(first); l = norm_name(last)
    if not f or not l:
        return None
    candidates = [
        ("first.last", f"{f}.{l}"),
        ("f.lastname", f"{f[:1]}.{l}"),
        ("firstname.l", f"{f}.{l[:1]}"),
        ("f.l", f"{f[:1]}.{l[:1]}"),
        ("firstlast", f"{f}{l}"),
        ("flast", f"{f[:1]}{l}"),
        ("lastfirst", f"{l}{f}"),
        ("last.f", f"{l}.{f[:1]}"),
        ("first", f"{f}"),
    ]
    for pat, tgt in candidates:
        if local == tgt:
            return pat
    return None

# -----------------------------
# Generate email from pattern
# -----------------------------
def generate_email(first, last, fmt, domain):
    if pd.isna(first) or pd.isna(last) or not domain:
        return None
    f = norm_name(first); l = norm_name(last)
    if not f or not l:
        return None
    fi, li = f[:1], l[:1]
    if   fmt == 'first.last':   local = f"{f}.{l}"
    elif fmt == 'f.lastname':   local = f"{fi}.{l}"
    elif fmt == 'firstname.l':  local = f"{f}.{li}"
    elif fmt == 'f.l':          local = f"{fi}.{li}"
    elif fmt == 'firstlast':    local = f"{f}{l}"
    elif fmt == 'flast':        local = f"{fi}{l}"
    elif fmt == 'lastfirst':    local = f"{l}{f}"
    elif fmt == 'last.f':       local = f"{l}.{fi}"
    elif fmt == 'first':        local = f"{f}"
    else: return None
    return f"{local}@{domain}"

# -----------------------------
# Repo loading
# -----------------------------
@st.cache_resource(show_spinner=False)
def load_repo_df():
    last_err = None
    for p in [p for p in CANDIDATE_REPO_PATHS if p]:
        try:
            if os.path.exists(p):
                df = pd.read_parquet(p)
                if sorted(df.columns.tolist()) != sorted(REPO_SCHEMA):
                    raise ValueError(
                        f"Wrong columns in {p}. Found {df.columns.tolist()}, expected {REPO_SCHEMA}."
                    )
                return df, p
        except Exception as e:
            last_err = e
    return None, f"No repository file found. Looked in: {CANDIDATE_REPO_PATHS}. Last error: {last_err}"

def build_maps_from_repo(repo_df: pd.DataFrame):
    cc = defaultdict(Counter)  # (company,country) -> Counter[(pattern,domain)]
    c  = defaultdict(Counter)  # company -> Counter[(pattern,domain)]
    for company, ctry, dom, pat, cnt in repo_df[["company","country_norm","domain","pattern","count"]].itertuples(index=False):
        cnt = int(cnt) if pd.notna(cnt) else 1
        cc[(company, ctry)][(pat, dom)] += cnt
        c[company][(pat, dom)] += cnt
    return cc, c

# -----------------------------
# Local inference maps (from uploaded file)
# -----------------------------
def build_maps_from_local(df, c_company, c_email, c_first, c_last, c_country):
    cc_local = defaultdict(Counter)
    c_local  = defaultdict(Counter)
    if not all([c_company, c_email, c_first, c_last]):
        return cc_local, c_local
    present = df[(df[c_email].notna()) & (df[c_email].astype(str).str.contains("@"))]
    for _, row in present.iterrows():
        comp = normalize_company_key(row[c_company])
        if not comp:
            continue
        email = str(row[c_email]).strip().lower()
        domain = email.split("@")[1] if "@" in email else None
        pat = detect_email_pattern(row[c_first], row[c_last], email)
        if not domain or not pat:
            continue
        ctry = normalize_country(row[c_country]) if c_country else None
        cc_local[(comp, ctry)][(pat, domain)] += 1
        c_local[comp][(pat, domain)] += 1
    return cc_local, c_local

# -----------------------------
# Helpers
# -----------------------------
def find_col(df, options):
    low = {c.lower(): c for c in df.columns}
    for o in options:
        if o.lower() in low:
            return low[o.lower()]
    return None

def detect_columns(df):
    c_company = find_col(df, ["Company","Company Name","Company Name for Emails","Account","Organisation"])
    c_email   = find_col(df, ["Email","Primary Email","Work Email","Business Email","E-mail"])
    c_first   = find_col(df, ["First Name","Firstname","Given Name","Forename"])
    c_last    = find_col(df, ["Last Name","Lastname","Surname","Family Name"])
    c_country = find_col(df, ["Country","Company Country","Office Country","HQ Country","Country/Region"])
    return c_company, c_email, c_first, c_last, c_country

def choose_best(counter: Counter):
    """Pick most common (pattern,domain). If tie, prefer .com."""
    if not counter:
        return None
    items = counter.most_common()
    best, best_count = items[0]
    ties = [x for x in items if x[1] == best_count]
    if len(ties) > 1:
        for (pat, dom), _ in ties:
            if dom.endswith(".com"):
                return (pat, dom)
    return best

# -----------------------------
# Fill: Local first, then Repo
# -----------------------------
def fill_hybrid(df, c_company, c_email, c_first, c_last, c_country,
                cc_local, c_local, cc_repo, c_repo):
    out = df.copy()
    filled_rows = []
    src_local_cc = src_local_c = src_repo_cc = src_repo_c = 0

    if not all([c_company, c_email, c_first, c_last]):
        raise ValueError("Need columns: Company, Email, First Name, Last Name.")

    target_idx = out[(out[c_email].isna()) | (out[c_email].astype(str).str.strip().eq(""))].index.tolist()
    total = len(target_idx) if target_idx else 1

    prog = st.progress(0)
    msg  = st.empty()

    for i, idx in enumerate(target_idx, 1):
        row = out.loc[idx]
        ck = normalize_company_key(row[c_company])
        if not ck:
            continue
        ct = normalize_country(row[c_country]) if c_country else None
        first, last = row[c_first], row[c_last]
        new_email = None

        # 1) LOCAL (company+country)
        best = choose_best(cc_local.get((ck, ct), Counter()))
        if best:
            pat, dom = best
            new_email = generate_email(first, last, pat, dom)
            if new_email: src_local_cc += 1

        # 2) LOCAL (company only)
        if new_email is None:
            best = choose_best(c_local.get(ck, Counter()))
            if best:
                pat, dom = best
                new_email = generate_email(first, last, pat, dom)
                if new_email: src_local_c += 1

        # 3) REPO (company+country)
        if new_email is None:
            best = choose_best(cc_repo.get((ck, ct), Counter()))
            if best:
                pat, dom = best
                new_email = generate_email(first, last, pat, dom)
                if new_email: src_repo_cc += 1

        # 4) REPO (company only)
        if new_email is None:
            best = choose_best(c_repo.get(ck, Counter()))
            if best:
                pat, dom = best
                new_email = generate_email(first, last, pat, dom)
                if new_email: src_repo_c += 1

        if new_email:
            out.at[idx, c_email] = new_email
            filled_rows.append(idx)

        if i % 25 == 0 or i == total:
            prog.progress(int(i / total * 100))
            msg.text(f"Processing {i}/{total}…")

    prog.progress(100)
    msg.empty()
    return out, filled_rows, {
        "from_local_company_country": src_local_cc,
        "from_local_company_only": src_local_c,
        "from_repo_company_country": src_repo_cc,
        "from_repo_company_only": src_repo_c,
        "considered": len(target_idx),
        "filled": len(filled_rows),
    }

# -----------------------------
# UI — Minimal
# -----------------------------
st.title("Email Filler")

# Tiny repo status line
repo_df, repo_info = load_repo_df()
col1, col2 = st.columns([3,2])
with col1:
    st.caption("Repository")
with col2:
    if repo_df is None:
        st.error("Not found")
    else:
        st.success("Loaded")

if repo_df is not None:
    st.caption(f"Path: {repo_info}  |  Rows: {len(repo_df):,}  |  Companies: {repo_df['company'].nunique():,}")

# Upload
upl = st.file_uploader("Upload CSV or Excel", type=["csv","xlsx","xls"])
if not upl:
    st.stop()

# Read input
if upl.name.lower().endswith(".csv"):
    df = pd.read_csv(upl, keep_default_na=True, low_memory=False)
else:
    df = pd.read_excel(upl, engine="openpyxl")

# Detect columns quietly
c_company, c_email, c_first, c_last, c_country = detect_columns(df)

# Gentle guardrails
needed = [("Company", c_company), ("Email", c_email), ("First Name", c_first), ("Last Name", c_last)]
missing = [name for name, col in needed if not col]
if missing:
    st.error(f"Missing required column(s): {', '.join(missing)}")
    st.stop()

# Build maps (local + repo)
cc_local, c_local = build_maps_from_local(df, c_company, c_email, c_first, c_last, c_country)
if repo_df is not None:
    cc_repo,  c_repo  = build_maps_from_repo(repo_df)
else:
    cc_repo, c_repo = defaultdict(Counter), defaultdict(Counter)

# Top-line quick facts
total_rows = len(df)
missing_emails = int((df[c_email].isna() | df[c_email].astype(str).str.strip().eq("")).sum())
st.write(f"**Rows:** {total_rows:,}   |   **Missing emails:** {missing_emails:,}")

# Run
if st.button("Run"):
    result_df, filled_idx, stats = fill_hybrid(
        df, c_company, c_email, c_first, c_last, c_country,
        cc_local, c_local, cc_repo, c_repo
    )

    # Compact summary panel
    remaining = missing_emails - stats["filled"]
    st.success(
        f"Filled {stats['filled']:,} of {stats['considered']:,} missing emails  •  "
        f"Local CC: {stats['from_local_company_country']:,}  |  "
        f"Local C: {stats['from_local_company_only']:,}  |  "
        f"Repo CC: {stats['from_repo_company_country']:,}  |  "
        f"Repo C: {stats['from_repo_company_only']:,}  |  "
        f"Left unfilled: {max(remaining, 0):,}"
    )

    # Downloads — Excel (highlight) + CSV
    import io
    from openpyxl.styles import PatternFill

    # mask for newly filled cells
    orig_missing_mask = (df[c_email].isna() | df[c_email].astype(str).str.strip().eq(""))
    filled_mask = result_df[c_email].notna() & orig_missing_mask

    # Excel with highlighting
    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Filled')
        ws = writer.sheets['Filled']
        email_col = list(result_df.columns).index(c_email) + 1
        yellow = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        for r, filled in enumerate(filled_mask, start=2):  # +1 header
            if filled:
                ws.cell(row=r, column=email_col).fill = yellow

    st.download_button(
        "Download Excel (highlighted)",
        data=excel_buf.getvalue(),
        file_name="emails_filled_highlighted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    csv_buf = io.BytesIO()
    result_df.to_csv(csv_buf, index=False)
    st.download_button("Download CSV", csv_buf.getvalue(), file_name="emails_filled.csv", mime="text/csv")
